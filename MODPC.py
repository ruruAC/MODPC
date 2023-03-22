from sklearn import datasets
import numpy as np
import random
import matplotlib.pyplot as plt
import time
import copy
import xlrd
import xlwt
import math
import pandas as pd
from  math import radians
from math import tan,atan,acos,sin,cos,asin,sqrt
import itertools
from sklearn.cluster import DBSCAN
import numpy as np
from openpyxl import load_workbook
import openpyxl
import requests
from sklearn import metrics
import gc
import joblib
import openpyxl

def geodistance(array_1,array_2):
    lng1 = array_1[0]
    lat1 = array_1[1]
    lng2 = array_2[0]
    lat2 = array_2[1]
    lng1, lat1, lng2, lat2 = map(radians, [lng1, lat1, lng2, lat2])
    dlon=lng2-lng1
    dlat=lat2-lat1
    a=sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    dis=2*asin(sqrt(a))*6371*1000
    #dis=sqrt((lng1-lng2)**2+(lat1-lat2)**2)
    return dis

def center(path):
    x,y=0,0
    for i in range(0,len(path)):
        x=path[i][0]+x
        y=path[i][1]+y
    x=x/len(path)
    y=y/len(path)
    return x,y
    
def midu(path,center):
    dis=[]
    for i in range(0,len(path)):
        p1=np.array(center)
        p2=np.array(path[i])
                   
        d=geodistance(p1,p2)
        dis.append(d)
                   
    m=max(dis)
    p=sum(dis)/len(path) #average dis
    s=0
    for i in range(0,len(path)):  #score
        if dis[i]<=p:
            s=s+1
    if p==0: 
        md=1
    else:
        md=(s*38*19)/(p**2*3.14)
    return p,md,m,dis

def judge(datamatrix,y_pred,r,flag):
     t=int(max(y_pred))
     if flag==2:
         noise.clear()
         newpath.clear()
         #print("xx",y_pred)
         flag=1
     for x in range(0,len(datamatrix)):
         if y_pred[x]==-1:
             c=[]
             c.append(datamatrix[x,0])
             c.append(datamatrix[x,1])
             noise.append(c)
     
     for i in range(0,t+1):
        path=[]  # 
        for x in range(0,len(datamatrix)):
            if y_pred[x]==i:
                c=[]
                c.append(datamatrix[x,0])
                c.append(datamatrix[x,1])
                path.append(c)
        
        x,y=center(path) 
        dj,aa,mda,julia=midu(path,[x,y])
        
        path=list(path)
        for item in noise:   
            x,y=center(path) 
            djk=geodistance([x,y],item)
            if djk<=500:      
                path.append(item)
                x,y=center(path) 
                dj,bb2,mda,julik=midu(path,[x,y])
               # print("<<<",djk,dk,bb1,bb2)
                if bb2<aa or dj>500 or mda>500: 
                    #print("xx",bb2,aa)

                    path.pop()  
                else:
                    noise.remove(item)
                    #print("yy",item,bb2,aa)
                    aa=bb2
        print("cut apart",dj,mda)             
        if dj>500 or mda>500:    
            db2(path,r+1,flag)
            
        else:
            newpath.append(path)
            #''path=tuple(path)
           
            #''write_excel2(pathj,path,j)
          
            
def judgenoise(path):
    x,y=center(path) 
    dj,aa,mda,julia=midu(path,[x,y])
    for item in noise:   
        x,y=center(path) 
        djk=geodistance([x,y],item)
        if djk<=500: 
            path.append(item)
            x,y=center(path)
            dj,bb2,mda,julik=midu(path,[x,y])
          
            if bb2<aa or mda>500 or dj>500: 
                path.pop() 
            else:
                noise.remove(item)
                    
                aa=bb2
    print("<<<<<<<<<")
    return path



            

                   

def db2(X,r,flag):
    
    datamatrix = np.zeros((len(X),2))
    for x in range(0,len(X)):
        datamatrix[x,0]=X[x][0]
        datamatrix[x,1]=X[x][1]  
    y_pred = DBSCAN(eps = 100, min_samples = r,metric=geodistance).fit_predict(datamatrix)
    judge(datamatrix,y_pred,r,flag)

             
def excel2m(path):
    data = xlrd.open_workbook(path)
    table = data.sheets()[0]
    nrows = table.nrows  
    ncols = table.ncols  
    c1 = np.arange(0, nrows, 1)
    datamatrix = np.zeros((nrows, ncols))
   
    for x in range(nrows):
        rows = table.row_values(x)
        datamatrix[x,0]=rows[1]
        datamatrix[x,1]=rows[2]
    return datamatrix
    del table
    gc.collect()

#写入初始汇聚结果
def write_excel(wjpath):
    wb=load_workbook(wjpath)
    wb1 = wb.active
    for i in range(1,len(X)+1):
        
        wb1.cell(i,1,y_pred[i-1])

    wb.save(wjpath)
    del wb1
    gc.collect


def write_path(wjpath):
    data = xlrd.open_workbook(wjpath)
    table = data.sheets()[0]
    nrows = table.nrows  
    ncols = table.ncols  
    t=table.col_values(0)
    global geshu
    geshu=int(max(t)) 
    global path
    path=[]
    for i in range(0,geshu+1):
        c=[]
        for x in range(nrows):
            rowsj=table.row_values(x)
            if rowsj[0]==i:
                
                datamatrix=[]
                data1=rowsj[1]
                data2=rowsj[2]
                datamatrix.append(data1)
                datamatrix.append(data2)
                c.append(datamatrix)
        path.append(c)
   # print(path)

                
    
def chaifen():
    global xxx
    xxx=geshu  #noise in new region
    for jj in range(0,geshu+1):
      
        if len(path[jj])==0:
            continue
        xj,yj=center(path[jj])
        dj,aa,mda,julia=midu(path[jj],[xj,yj])
 
        if dj>500 or mda>500 : 
            global newpath #save the remaining region
            newpath=[]   
            global noise
            noise=[]
            print("----cut apart----",jj)
            db2(path[jj],5,1) #4+1
            for c in newpath:  
               #print(c)
               c=judgenoise(c)
               path.append(c)
               xxx=xxx+1
               print("----Increase---",xxx)

            path[jj]=[] 
            if len(noise)==0:
                continue
            db2(noise,4,2) #recognize the noise
            for c in newpath:  
               #print(c)
               c=judgenoise(c) 
               path.append(c)
               xxx=xxx+1
               print("----Increase---",xxx)
    
def hebing(): #Traverse each region for merging
    for jj in range(0,xxx+1):
      
        if len(path[jj])==0:
            continue
        xj,yj=center(path[jj])
        dj,aa,mda,julia=midu(path[jj],[xj,yj])
        #print(jj)
        #print(dj,aa,mda)
        for kk in range(0,xxx+1):
           
            if len(path[jj])==0:
                break
            if len(path[kk])==0:
                continue
            if kk==jj:
                continue
            xk,yk=center(path[kk]) 
            
            dk,bb,mdb,julib=midu(path[kk],[xk,yk])
            if geodistance([xk,yk],[xj,yj])>1000:
                continue
       
            pathjk=[]
            for item in path[jj]:
                pathjk.append(item)
            for item in path[kk]:
                pathjk.append(item)
            xjk=(xj*len(path[jj])+(xk*len(path[kk])))/(len(path[jj])+len(path[kk]))
            yjk=(yj*len(path[jj])+(yk*len(path[kk])))/(len(path[jj])+len(path[kk]))
         
            djk,cc,mdc,julic=midu(pathjk,[xjk,yjk]) 
            ddd=geodistance([xj,yj],[xk,yk])
            if (dj+dk)==0:
                sss=1
            else:
                sss=ddd/(dj+dk) 
            
            #print(sss)
            #print(">>>>>>>>")
            if djk<=500 and mdc<=500 and sss<=2: # merge or not 

                print("--merge--",jj,"->",kk)
                path[jj]=[]
                path[kk]=pathjk

def neisuo(jj):
    
    xj,yj=center(path[jj]) 
    dj,aa1,md,juli=midu(path[jj],[xj,yj])
    pathjj=list(path[jj])
    if dj==0:
        print("dj=0",jj)
        return
    while md/dj>=1.8 :
        
        m=juli.index(max(juli))
        mm=max(juli)
        item=pathjj[m]
        pathjj.remove(item)
        xj,yj=center(pathjj)
        dj,aa1,md,juli=midu(pathjj,[xj,yj]) 

        for kk in range(0,xxx+1):
            flag=0
           
            if kk==jj:
                continue
            if len(path[kk])==0:
                continue
            xk,yk=center(path[kk]) 
            dk,bb1,mdb,julik=midu(path[kk],[xk,yk])
            pathkk=list(path[kk])

            djk=geodistance([xk,yk],item) 
            if djk<=500:    
                pathkk.append(item)
                xk,yk=center(pathkk) #the center of cluster k 
                dk2,bb2,mdbh,julik=midu(pathkk,[xk,yk]) 
               
                #print("<<<",djk,dk,bb1,bb2)
                if bb2<bb1 or dk2>500 or mdbh>500: 
                    pathkk.pop()  #remove
                else:

                    bb1=bb2
                    flag=1
                  
            if flag==1: 
                path[kk]=pathkk
                break
        
        if dj==0:
            break
    print("----limit----",jj)

    path[jj]=pathjj 
   

def dianjuhe(jj):
    for kk in range(0,xxx+1):
        flag=0
        if kk==jj:
            continue
       
        if len(path[kk])==0:
            continue
        if len(path[jj])==0:
            break
      
        #print(pathkk)
        xk,yk=center(path[kk]) 
        xj,yj=center(path[jj]) #the center of cluster k 
        dk,bb1,mdb,julik=midu(path[kk],[xk,yk])
       
        dj,aa1,mda,julia=midu(path[jj],[xj,yj])
        dcen=geodistance([xk,yk],[xj,yj])
        
        if dcen>=1000:
            continue

        for item in path[jj]:
            xk,yk=center(path[kk]) 
            djk=geodistance([xk,yk],item) 
            if djk<=500:      
                path[kk].append(item)
                xk,yk=center(path[kk]) #the center of cluster k 
                dk2,bb2,mdbh,julik=midu(path[kk],[xk,yk]) 
                xj,yj=center(path[jj]) #the center of cluster k 
                dj2,aa2,mda,julia=midu(path[jj],[xj,yj])
                if bb2<bb1 or dk2>500 or mdbh>500 or (aa1>=0.3 and (aa1>aa2 or mda>500 or dj2>500)): #把点合进簇k后，导致K的密度降低 撤回 aa1>0.3说明 jj是要保留的，如果拆掉元素对他也有影响需要撤回
                    path[kk].pop()  
                else:
                    aa1=aa2
                    bb1=bb2
                    flag=1
                    path[jj].remove(item)  
        if flag==1: 
            print("----transform----")
            print(jj,"->",kk)
           
           
def sicijulei():
    for jj in range(6734,xxx+1):
        if len(path[jj])==0:
            continue
        xj,yj=center(path[jj])
        dj,aa,mda,julij=midu(path[jj],[xj,yj]) #dj: distance aa: density mda: the farest juli:list
        print(jj,dj,aa,mda)
        neisuo(jj)
        if dj<150:
            xj,yj=center(path[jj])
            dj,aa,mda,julij=midu(path[jj],[xj,yj])
            if aa<0.2 or len(path[jj])<4:
                print("--ignore--",jj)
                dianjuhe(jj)
                path[jj]=[]
                

def write_excel_quyu_last(quyupath):
    wb=load_workbook(quyupath)
    wb1 = wb.active
    h=1
    for jj in range(0,xxx+1):
        if len(path[jj])==0:
            continue
        xj,yj=center(path[jj])
        dj,aa,mda,julij=midu(path[jj],[xj,yj])
        for x in path[jj]:
            
            wb1.cell(h,1,jj)
            wb1.cell(h,2,x[0])
            wb1.cell(h,3,x[1])
            wb1.cell(h,4,dj)
            wb1.cell(h,5,aa)
            wb1.cell(h,6,mda)
            h=h+1
    wb.save(quyupath)
    del wb1
    gc.collect()


 
wjpath='郑州弱覆盖.xlsx'
quyupath='郑州弱覆盖2.xlsx'
X = excel2m(wjpath)
y_pred = DBSCAN(eps = 100, min_samples = 4,metric=geodistance).fit_predict(X)
write_excel(wjpath)
write_path(wjpath)
chaifen()
#path=joblib.load('path_si.pkl') 
#xxx=10761
hebing()
joblib.dump(path,'path_1.pkl')
sicijulei()
joblib.dump(path,'path_2.pkl')
write_excel_quyu_last(quyupath)

